#!/usr/bin/env python3
# /// script
# requires-python = ">=3.10"
# dependencies = [
#     "instaloader>=4.13.0",
#     "openpyxl>=3.1.5",
# ]
# ///
"""
Instagram Scraper for Competitive Intelligence
Extracts all parameters from public Instagram profiles and saves to Excel
Uses instaloader for reliable data extraction
"""

import re
import instaloader
import openpyxl
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment

class InstaLoaderScraper:
    """
    Instagram scraper using instaloader library
    Reliable and handles Instagram's current structure
    """
    
    def __init__(self, username):
        self.username = username
        self.posts_data = []
        self.profile_data = {}
        self.loader = instaloader.Instaloader(
            quiet=False,
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        )
    
    def scrape(self):
        """Main scraping method"""
        print(f"🔍 Scraping @{self.username}...")
        
        try:
            # Fetch profile
            print(f"📍 Fetching profile data...")
            profile = instaloader.Profile.from_username(self.loader.context, self.username)
            
            # Extract profile info
            self.profile_data = {
                'username': self.username,
                'followers': profile.followers,
                'following': profile.followees,
                'bio': profile.biography,
                'is_business': profile.is_business_account,
                'is_verified': profile.is_verified,
                'total_posts': profile.mediacount,
                'profile_url': f"https://www.instagram.com/{self.username}/"
            }
            
            print(f"✅ Profile: @{self.username}")
            print(f"   Followers: {self.profile_data['followers']:,}")
            print(f"   Following: {self.profile_data['following']:,}")
            print(f"   Posts: {self.profile_data['total_posts']}")
            print(f"   Verified: {self.profile_data['is_verified']}")
            
            # Extract posts
            print(f"📸 Fetching posts...")
            posts = profile.get_posts()
            
            for idx, post in enumerate(posts, 1):
                if idx > 15:  # Get last 15 posts
                    break
                self.extract_post_details(post, idx)
            
            print(f"✅ Extracted {len(self.posts_data)} posts")
            return True
            
        except Exception as e:
            print(f"❌ Error: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
    
    def extract_post_details(self, post, position):
        """Extract individual post details"""
        
        try:
            caption = post.caption if post.caption else 'N/A'
            
            # Extract hashtags
            hashtags = re.findall(r'#\w+', caption)
            tags = ', '.join(hashtags) if hashtags else 'N/A'
            
            # Determine content type
            if post.is_video:
                if hasattr(post, 'is_carousel') and post.is_carousel:
                    content_type = 'Carousel'
                else:
                    content_type = 'Video/Reel'
            else:
                if hasattr(post, 'is_carousel') and post.is_carousel:
                    content_type = 'Carousel'
                else:
                    content_type = 'Image'
            
            # Get date
            date_posted = post.date_utc.strftime('%Y-%m-%d %H:%M:%S')
            
            # Extract media URLs
            photos = self.extract_media_urls(post, 'image')
            videos = self.extract_media_urls(post, 'video')
            
            post_data = {
                'Position': position,
                'Post URL': f"https://www.instagram.com/p/{post.shortcode}/",
                'Cover Image': post.url,
                'Description': caption[:100] + '...' if len(caption) > 100 else caption,
                'Posted User Profile': self.profile_data['profile_url'],
                'Content Type': content_type,
                'Tags': tags,
                'Date Posted': date_posted,
                'Number of Likes': post.likes,
                'Photos': photos,
                'Videos': videos,
                'Video View Count': post.video_view_count if post.is_video else 'N/A',
                'Video Play Count': 'N/A',  # Not available in instaloader
                'Number of Comments': post.comments,
                'Latest Comments': self.extract_comments(post),
                'Followers': self.profile_data['followers'],
                'Engagement Rate': f"{((post.likes + post.comments) / self.profile_data['followers'] * 100):.2f}%" if self.profile_data['followers'] > 0 else 'N/A',
            }
            
            self.posts_data.append(post_data)
            print(f"  ✓ Post {position}: {post_data['Date Posted']} ({content_type}) - {post.likes} likes")
            
        except Exception as e:
            print(f"  ⚠️  Error extracting post {position}: {str(e)}")
    
    def extract_media_urls(self, post, media_type='image'):
        """Extract image or video URLs"""
        urls = []
        
        try:
            if media_type == 'image':
                # Get display URL (works for images and carousels)
                if hasattr(post, 'url') and post.url:
                    urls.append(post.url)
            elif media_type == 'video' and post.is_video:
                # Get video URL
                if hasattr(post, 'video_url') and post.video_url:
                    urls.append(post.video_url)
        except Exception as e:
            pass
        
        return ' | '.join(urls) if urls else 'N/A'
    
    def extract_comments(self, post):
        """Extract latest comments"""
        comments = []
        
        try:
            for comment in post.get_comments():
                text = comment.text
                author = comment.owner.username
                if text and author:
                    comments.append(f"{author}: {text[:40]}...")
                if len(comments) >= 2:
                    break
        except:
            pass
        
        return ' | '.join(comments) if comments else 'N/A'
    
    def save_excel(self, filename=None):
        """Save to Excel file"""
        if filename is None:
            filename = f"{self.username}_instagram_data.xlsx"
        
        if not self.posts_data:
            print("❌ No data to save!")
            return False
        
        print(f"\n💾 Saving to {filename}...")
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Posts"
        
        # Get columns
        columns = list(self.posts_data[0].keys())
        
        # Write header
        for col_idx, column in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=column)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="1F1F1F", end_color="1F1F1F", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Write data
        for row_idx, post in enumerate(self.posts_data, 2):
            for col_idx, column in enumerate(columns, 1):
                value = post.get(column, 'N/A')
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.font = Font(size=10)
        
        # Set column widths
        widths = {
            'A': 8, 'B': 30, 'C': 20, 'D': 25, 'E': 25, 'F': 15,
            'G': 25, 'H': 18, 'I': 12, 'J': 25, 'K': 25, 'L': 15,
            'M': 15, 'N': 30, 'O': 12, 'P': 15
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        
        # Freeze header
        ws.freeze_panes = 'A2'
        
        # Save
        wb.save(filename)
        print(f"✅ Saved: {filename}")
        print(f"📊 Total posts: {len(self.posts_data)}")
        
        return True


def main():
    """Main execution"""
    
    # Configure this:
    username = "ikfdigital"  # Change this to any Instagram handle
    
    print("=" * 60)
    print("📱 Instagram Competitor Scraper (Powered by Instaloader)")
    print("=" * 60)
    
    # Run scraper
    scraper = InstaLoaderScraper(username)
    
    if scraper.scrape():
        scraper.save_excel()
        print("\n✅ Done! Excel file ready.")
    else:
        print("\n❌ Scraping failed. Check your internet connection and username.")


if __name__ == "__main__":
    main()
