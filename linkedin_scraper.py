#!/usr/bin/env python3
# /// script
# requires-python = ">=3.10"
# dependencies = [
#     "selenium>=4.0.0",
#     "webdriver-manager>=4.0.0",
#     "openpyxl>=3.1.5",
# ]
# ///
"""
LinkedIn Multi-Profile Scraper (Headless Selenium)
Runs completely in background - no visible browser window
Uses real browser automation to bypass LinkedIn security

No cookies needed - real browser login!
"""

import openpyxl
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
import time
import re


class LinkedInMultiProfileScraper:
    """
    Multi-profile LinkedIn scraper using Selenium (headless)
    Real browser automation - completely in background, no visible window
    """
    
    def __init__(self, email: str, password: str):
        self.email = email
        self.password = password
        self.driver = None
        self.all_data = []
    
    def convert_relative_date_to_actual(self, relative_date: str) -> str:
        """Convert LinkedIn relative dates (2d, 3w, 1mo) to actual dates"""
        try:
            from datetime import timedelta

            # Defensive: handle None/float/etc. gracefully
            if relative_date is None:
                return "N/A"
            if not isinstance(relative_date, str):
                return str(relative_date)

            raw = relative_date

            # Normalize and strip noisy tokens LinkedIn often includes
            cleaned = raw.lower()
            cleaned = cleaned.replace('edited', '')
            cleaned = cleaned.replace('ago', '')
            cleaned = cleaned.replace('•', ' ').replace('·', ' ')
            cleaned = re.sub(r'\s+', ' ', cleaned).strip()

            # Handle immediate times
            if cleaned in ['just now', 'now']:
                return datetime.now().strftime('%Y-%m-%d')
            if 'today' in cleaned:
                return datetime.now().strftime('%Y-%m-%d')
            if 'yesterday' in cleaned:
                return (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')

            # Try absolute date formats first (e.g., "Dec 5", "December 5, 2024")
            absolute_formats = [
                '%b %d, %Y', '%B %d, %Y', '%b %d %Y', '%B %d %Y',
                '%b %d', '%B %d',
            ]
            for fmt in absolute_formats:
                try:
                    parsed = datetime.strptime(cleaned.title(), fmt)
                    # If year missing, assume current year and adjust if future
                    if '%Y' not in fmt:
                        parsed = parsed.replace(year=datetime.now().year)
                        if parsed > datetime.now():
                            parsed = parsed.replace(year=parsed.year - 1)
                    return parsed.strftime('%Y-%m-%d')
                except ValueError:
                    pass

            # Relative formats like "3w", "2 weeks", "1 yr"
            match = re.search(r'(\d+)\s*(h|hr|hrs|hour|hours|d|day|days|w|wk|wks|week|weeks|m|mo|mos|month|months|y|yr|yrs|year|years)', cleaned)
            if not match:
                return raw  # Return original if we cannot parse

            number = int(match.group(1))
            unit = match.group(2)

            if unit.startswith('h'):
                actual_date = datetime.now() - timedelta(hours=number)
            elif unit.startswith('d'):
                actual_date = datetime.now() - timedelta(days=number)
            elif unit.startswith('w'):
                actual_date = datetime.now() - timedelta(weeks=number)
            elif unit.startswith('m'):
                actual_date = datetime.now() - timedelta(days=number * 30)
            elif unit.startswith('y'):
                actual_date = datetime.now() - timedelta(days=number * 365)
            else:
                return raw

            return actual_date.strftime('%Y-%m-%d')
        except Exception:
            return str(relative_date)

    def extract_hashtags(self, content_text: str = "", html_text: str = None, extra_texts: list | None = None) -> list:
        """Collect hashtags from text plus HTML (links/encoded)."""
        texts = []
        if content_text:
            texts.append(content_text)
        if extra_texts:
            texts.extend([t for t in extra_texts if t])
        if html_text:
            texts.append(html_text)

        tags = []
        seen = set()

        # Regex for visible hashtags; allows letters, numbers, underscores; length guard avoids noise
        tag_pattern = re.compile(r'#([A-Za-z0-9_]{2,60})')

        for txt in texts:
            if not isinstance(txt, str):
                continue
            # Separate concatenated hashtags like "#AI#ML" before scanning
            normalized = re.sub(r'(?<!\s)#', ' #', txt)
            for match in tag_pattern.findall(normalized):
                candidate = f"#{match}"
                key = candidate.lower()
                if key not in seen:
                    seen.add(key)
                    tags.append(candidate)

        if html_text:
            # Capture hashtags present in links such as /hashtag/ai or %23ai
            for match in re.findall(r'/hashtag/([A-Za-z0-9_-]{2,60})', html_text, re.IGNORECASE):
                candidate = f"#{match}"
                key = candidate.lower()
                if key not in seen:
                    seen.add(key)
                    tags.append(candidate)
            for match in re.findall(r'%23([A-Za-z0-9_-]{2,60})', html_text, re.IGNORECASE):
                candidate = f"#{match}"
                key = candidate.lower()
                if key not in seen:
                    seen.add(key)
                    tags.append(candidate)

        return tags
    
    def extract_username_from_url(self, url_or_username: str) -> str:
        """Extract username from LinkedIn URL or return as-is if already username"""
        url = url_or_username.strip()
        
        # If it's already just a username (no slashes), return it
        if '/' not in url:
            return url
        
        # Extract username from various LinkedIn URL formats
        # Personal: https://www.linkedin.com/in/username/
        # Company: https://www.linkedin.com/company/company-name/
        
        try:
            # Remove protocol if present
            url = url.replace('https://', '').replace('http://', '')
            
            # Remove www. if present
            url = url.replace('www.', '')
            
            # Check if it's a company page
            if '/company/' in url:
                company_name = url.split('/company/')[1].split('/')[0].split('?')[0]
                return company_name
            
            # Extract username from personal profile
            elif '/in/' in url:
                username = url.split('/in/')[1].split('/')[0].split('?')[0]
                return username
            else:
                # If no /in/ or /company/ found, maybe it's just the username/company name
                return url.split('/')[0]
        except:
            # If parsing fails, return original
            return url_or_username
    
    def setup_driver(self):
        """Setup headless Chrome driver - runs in background"""
        print("🌐 Setting up browser (headless - no window)...")
        
        options = Options()
        
        # CRITICAL: Make browser completely headless (invisible)
        options.add_argument('--headless=new')  # Newer headless mode
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--disable-blink-features=AutomationControlled')
        
        # Realistic user agent
        options.add_argument('user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36')
        
        # Window size for headless mode
        options.add_argument('--window-size=1920,1080')
        
        service = Service(ChromeDriverManager().install())
        self.driver = webdriver.Chrome(service=service, options=options)
        print("✅ Browser ready (running headless in background)")
        
    def login(self):
        """Login to LinkedIn"""
        print(f"🔐 Logging in as {self.email}...")
        
        try:
            self.driver.get("https://www.linkedin.com/login")
            time.sleep(2)
            
            # Enter email
            email_field = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.ID, "username"))
            )
            email_field.send_keys(self.email)
            
            # Enter password
            password_field = self.driver.find_element(By.ID, "password")
            password_field.send_keys(self.password)
            
            # Click login
            login_button = self.driver.find_element(By.CSS_SELECTOR, "button[type='submit']")
            login_button.click()
            
            # Wait for redirect
            time.sleep(5)
            
            # Check if logged in
            if "/feed" in self.driver.current_url or "/in/" in self.driver.current_url:
                print("✅ Login successful!")
                return True
            else:
                print(f"⚠️  Current URL: {self.driver.current_url}")
                print("   Waiting for page to load...")
                time.sleep(10)
                
                if "/feed" in self.driver.current_url or "/in/" in self.driver.current_url:
                    print("✅ Login successful!")
                    return True
                else:
                    print("❌ Login may have failed - proceeding anyway")
                    return True  # Proceed anyway
                
        except Exception as e:
            print(f"❌ Login error: {str(e)}")
            return False
    
    def scrape_profile(self, profile_url_or_username: str, max_posts: int = 15):
        """Scrape a LinkedIn profile (personal or company) from URL or username"""
        # Extract username/company name from URL
        profile_name = self.extract_username_from_url(profile_url_or_username)
        
        # Determine if it's a company or personal profile
        is_company = '/company/' in profile_url_or_username
        profile_type = "Company" if is_company else "Personal"
        
        print(f"\n{'='*70}")
        print(f"📍 Scraping {profile_type} profile: {profile_name}")
        print(f"   Original input: {profile_url_or_username}")
        print(f"{'='*70}")
        
        try:
            # Navigate to profile - handle both company and personal profiles
            if is_company:
                # For company pages, go directly to posts section
                print("🎯 Navigating directly to company posts section...")
                posts_url = f"https://www.linkedin.com/company/{profile_name}/posts/"
                self.driver.get(posts_url)
                time.sleep(5)
            else:
                # For personal profiles, go to main profile first
                profile_url = f"https://www.linkedin.com/in/{profile_name}/"
                self.driver.get(profile_url)
                time.sleep(5)
            
            # Save screenshot for debugging
            try:
                self.driver.save_screenshot(f"debug_{profile_name}_profile.png")
                print(f"📸 Screenshot saved: debug_{profile_name}_profile.png")
            except:
                pass
            
            # Extract profile info
            profile_data = self.extract_profile_info(profile_name, is_company=is_company)
            
            # For company pages, we're already at posts - just scroll
            # For personal profiles, we need to navigate to posts if few are found
            if is_company:
                print("📜 Scrolling company posts page to load all posts...")
                self.scroll_page(scrolls=12)  # More aggressive scrolling for company posts
                posts_extracted = self.extract_posts(profile_data, max_posts)
            else:
                print("📜 Scrolling main profile page to find posts...")
                self.scroll_page(scrolls=8)
                posts_extracted = self.extract_posts(profile_data, max_posts)
                
                # If few posts found, try activity page
                if posts_extracted < 5:
                    print(f"\n⏳ Found only {posts_extracted} posts. Trying activity page...")
                    posts_url = f"https://www.linkedin.com/in/{profile_name}/recent-activity/all/"
                    self.driver.get(posts_url)
                    time.sleep(5)

                
                # Save screenshot for debugging
                try:
                    self.driver.save_screenshot(f"debug_{profile_name}_activity.png")
                    print(f"📸 Screenshot saved: debug_{profile_name}_activity.png")
                except:
                    pass
                
                # Scroll to load posts
                print("📜 Scrolling activity page...")
                self.scroll_page(scrolls=5)
                
                # Extract from activity page
                posts_extracted = self.extract_posts(profile_data, max_posts)
            
            print(f"✅ Extracted {posts_extracted} posts from @{profile_name}")
            return True
            
        except Exception as e:
            print(f"❌ Error scraping @{profile_name}: {str(e)}")
            return False
    
    def extract_profile_info(self, profile_name: str, is_company: bool = False) -> dict:
        """Extract profile information from page (personal or company)"""
        
        try:
            # Get profile name
            try:
                name_elem = self.driver.find_element(By.CSS_SELECTOR, "h1")
                full_name = name_elem.text
            except:
                full_name = "N/A"
            
            # Get headline
            try:
                headline_elems = self.driver.find_elements(By.CSS_SELECTOR, "div.text-body-medium")
                headline = headline_elems[0].text if headline_elems else "N/A"
            except:
                headline = "N/A"
            
            # Get location
            try:
                location_elem = self.driver.find_element(By.CSS_SELECTOR, "span.text-body-small")
                location = location_elem.text
            except:
                location = "N/A"
            
            # Get follower count - try multiple methods
            followers = "N/A"
            try:
                # Method 1: Look for "followers" text
                page_text = self.driver.find_element(By.TAG_NAME, "body").text
                
                # Extract followers number (e.g., "1,234 followers")
                follower_match = re.search(r'([\d,]+)\s*followers?', page_text, re.IGNORECASE)
                if follower_match:
                    followers = int(follower_match.group(1).replace(',', ''))
                else:
                    # Method 2: Try looking for specific elements
                    follower_elems = self.driver.find_elements(By.XPATH, "//*[contains(text(), 'follower')]")
                    if follower_elems:
                        for elem in follower_elems:
                            text = elem.text
                            numbers = re.findall(r'[\d,]+', text)
                            if numbers:
                                followers = int(numbers[0].replace(',', ''))
                                break
            except Exception as e:
                print(f"    ⚠️  Could not extract followers: {str(e)}")
            
            # Get connections count
            connections = "N/A"
            try:
                page_text = self.driver.find_element(By.TAG_NAME, "body").text
                
                # Extract connections number
                conn_match = re.search(r'([\d,]+)\s*(?:contacts|connections?)', page_text, re.IGNORECASE)
                if conn_match:
                    connections = int(conn_match.group(1).replace(',', ''))
            except:
                pass
            
            # Get profile summary (bio/about)
            summary = "N/A"
            try:
                page_text = self.driver.find_element(By.TAG_NAME, "body").text
                
                # Look for "About" section
                about_match = re.search(r'About\s*([^\n]+(?:\n[^\n]+){0,3})', page_text)
                if about_match:
                    summary = about_match.group(1).strip()
                    if len(summary) > 500:
                        summary = summary[:500]
                else:
                    # Try alternative: look for any text after headline
                    lines = page_text.split('\n')
                    for i, line in enumerate(lines):
                        if 'headline' in line.lower() or full_name in line:
                            # Get next non-empty lines as summary
                            for j in range(i+1, min(i+5, len(lines))):
                                if lines[j].strip() and not any(keyword in lines[j].lower() for keyword in ['followers', 'connections', 'contact', 'profile']):
                                    summary = lines[j].strip()
                                    break
                            break
            except:
                pass
            
            # Build profile URL based on type
            if is_company:
                profile_url = f"https://www.linkedin.com/company/{profile_name}/"
            else:
                profile_url = f"https://www.linkedin.com/in/{profile_name}/"
            
            profile_data = {
                'profile_username': profile_name,
                'profile_full_name': full_name,
                'profile_headline': headline[:100] if len(headline) > 100 else headline,
                'profile_location': location,
                'profile_url': profile_url,
                'profile_followers': followers,
                'profile_connections': connections,
                'profile_summary': summary,
                'profile_type': 'Company' if is_company else 'Personal'
            }
            
            print(f"✅ Profile: {full_name}")
            print(f"   Headline: {headline[:60]}...")
            print(f"   Location: {location}")
            print(f"   Followers: {followers}")
            print(f"   Connections: {connections}")
            print(f"   Summary: {summary[:60] if summary != 'N/A' else 'N/A'}...")
            
            return profile_data
            
        except Exception as e:
            print(f"⚠️  Error extracting profile: {str(e)}")
            return {
                'profile_username': profile_username,
                'profile_full_name': 'N/A',
                'profile_headline': 'N/A',
                'profile_location': 'N/A',
                'profile_url': f"https://www.linkedin.com/in/{profile_username}/",
                'profile_followers': 'N/A',
                'profile_connections': 'N/A',
                'profile_summary': 'N/A'
            }
    
    def scroll_page(self, scrolls: int = 5):
        """Scroll page to load more posts"""
        for i in range(scrolls):
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(2)
            print(f"  Scroll {i+1}/{scrolls}")
    
    def extract_posts(self, profile_data: dict, max_posts: int) -> int:
        """Extract posts from page"""
        
        print("📝 Extracting posts...")
        posts_extracted = 0
        
        try:
            # Find all post containers
            post_selectors = [
                "div[data-feed-item-type]",  # General feed items
                "div.feed-shared-update-v2",
                "article",
                "div.base-card",
            ]
            
            post_elements = []
            selector_used = ""
            
            for selector in post_selectors:
                try:
                    elements = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    if elements and len(elements) > len(post_elements):
                        post_elements = elements
                        selector_used = selector
                except:
                    pass
            
            if selector_used:
                print(f"  Using selector: {selector_used} (found {len(post_elements)} elements)")
            
            # Remove duplicate elements using ROBUST deduplication
            # Track by combining multiple attributes: content hash + date + position
            unique_posts = []
            seen_fingerprints = set()
            
            for elem in post_elements:
                try:
                    # Get full text content
                    post_text = elem.text.strip()
                    
                    # Extract key identifiers for deduplication
                    # 1. Content hash (first 150 chars + last 50 chars)
                    content_signature = post_text[:150] + "|||" + post_text[-50:] if len(post_text) > 200 else post_text
                    
                    # 2. Try to extract date/time from post
                    date_match = re.search(r'(\d+[hdwmo]|today|yesterday)', post_text, re.IGNORECASE)
                    date_sig = date_match.group(1) if date_match else "unknown"
                    
                    # 3. Get element position
                    try:
                        location = elem.location
                        # Use actual Y coordinate (not grouped)
                        y_pos = location.get('y', 0)
                    except:
                        y_pos = 0
                    
                    # 4. Create comprehensive fingerprint
                    fingerprint = (content_signature, date_sig, y_pos)
                    
                    # Check if we've seen this exact post before
                    if fingerprint not in seen_fingerprints:
                        unique_posts.append(elem)
                        seen_fingerprints.add(fingerprint)
                    else:
                        print(f"    [DUPLICATE SKIPPED] Same post detected")
                        
                except Exception as e:
                    # If error, still include the element
                    unique_posts.append(elem)
            
            # Limit to max posts
            post_elements = unique_posts[:max_posts]
            
            print(f"📊 Found {len(post_elements)} UNIQUE post elements (after dedup)")
            
            # If no structured posts found, try to get text content
            if not post_elements:
                print("⚠️  No structured posts found, trying alternative extraction...")
                
                # Try scrolling and getting all text
                page_text = self.driver.find_element(By.TAG_NAME, "body").text
                
                # Simple heuristic: split by timestamps or post indicators
                lines = page_text.split('\n')
                
                post_count = 0
                for i, line in enumerate(lines):
                    if ('ago' in line.lower() or 'week' in line.lower() or 'day' in line.lower()) and i > 0:
                        # Found a potential post
                        post_text = lines[i-1] if i > 0 else "N/A"
                        
                        # Extract hashtags from post text
                        hashtags = re.findall(r'#[A-Za-z0-9_]+', post_text)
                        hashtags_str = ','.join(hashtags) if hashtags else 'None'
                        
                        # Convert date
                        actual_date = self.convert_relative_date_to_actual(line)
                        
                        post_data = {
                            'Profile Username': profile_data['profile_username'],
                            'Profile Name': profile_data['profile_full_name'],
                            'Profile Headline': profile_data['profile_headline'],
                            'Profile Location': profile_data['profile_location'],
                            'Profile URL': profile_data['profile_url'],
                            'Profile Followers': profile_data['profile_followers'],
                            'Profile Summary': profile_data['profile_summary'],
                            'Post Position': post_count + 1,
                            'Post URL': 'N/A',
                            'Post ID': 'N/A',
                            'Post Date': actual_date,
                            'Post Content (Full)': post_text,
                            'Content URL': 'N/A',
                            'Post Word Count': len(post_text.split()),
                            'Content Type': 'Text Only',
                            'Has Document': 'No',
                            'Hashtags': hashtags_str,
                            'Hashtag Count': len(hashtags),
                            'Post Reactions (Likes)': 0,
                            'Post Comments': 0,
                            'Post Shares': 0,
                            'Total Engagement': 0,
                            'Engagement Rate (%)': '0.00%',
                            'Comment/Reaction Ratio': '0',
                            'Share/Reaction Ratio': '0%',
                            'Scraped At': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        }
                        
                        self.all_data.append(post_data)
                        posts_extracted += 1
                        post_count += 1
                        
                        if post_count >= max_posts:
                            break
                
                return posts_extracted
            
            for idx, post_elem in enumerate(post_elements, 1):
                try:
                    post_data = self.extract_single_post(post_elem, profile_data, idx)
                    if post_data:
                        self.all_data.append(post_data)
                        posts_extracted += 1
                        print(f"  ✓ Post {idx}: {post_data.get('Post Date', 'N/A')}")
                except Exception as e:
                    print(f"  ⚠️  Error extracting post {idx}: {str(e)}")
                    continue
            
            return posts_extracted
            
        except Exception as e:
            print(f"❌ Error extracting posts: {str(e)}")
            return 0
    
    def extract_single_post(self, post_elem, profile_data: dict, position: int) -> dict:
        """Extract data from single post element"""
        
        try:
            # Extract text - try multiple selectors
            content_text = "N/A"
            content_html = None  # Keep the HTML only for the text block to avoid grabbing page-level hashtags
            text_selectors = [
                "span.break-words",
                "div.feed-shared-text",
                "div[dir='ltr']",
                ".feed-shared-update-v2__commentary",
                ".feed-shared-text__text-view",
                "span[dir='ltr']",
            ]
            
            for selector in text_selectors:
                try:
                    text_elem = post_elem.find_element(By.CSS_SELECTOR, selector)
                    if text_elem and text_elem.text.strip():
                        content_text = text_elem.text.strip()
                        try:
                            content_html = text_elem.get_attribute("innerHTML")
                        except Exception:
                            content_html = None
                        break
                except:
                    continue
            
            # If still no text, try getting all visible text from post
            if content_text == "N/A" or len(content_text) < 10:
                try:
                    content_text = post_elem.text.strip()
                    # Remove first few lines that are usually name/time
                    lines = content_text.split('\n')
                    if len(lines) > 3:
                        content_text = '\n'.join(lines[3:])
                except:
                    pass
            
            # Extract date - try multiple selectors and methods
            date_text = "N/A"
            date_selectors = [
                "span.feed-shared-actor__sub-description",
                "time",
                ".feed-shared-actor__sub-description",
                "span[class*='time']",
                ".update-components-actor__sub-description",
            ]
            
            for selector in date_selectors:
                try:
                    date_elem = post_elem.find_element(By.CSS_SELECTOR, selector)
                    if date_elem and date_elem.text.strip():
                        date_text = date_elem.text.strip()
                        break
                except:
                    continue
            
            # If still no date, try extracting from post text more aggressively
            if date_text == "N/A" or len(date_text) < 2:
                try:
                    post_text = post_elem.text
                    
                    # Pattern 1: Look for time indicators like "2d", "1w", "3mo", "Today", "Yesterday"
                    time_patterns = re.findall(r'(?:today|yesterday|\d+[hdwmo])(?:\s*\u2022)?', post_text, re.IGNORECASE)
                    if time_patterns:
                        date_text = time_patterns[0].strip('•').strip()
                    
                    # Pattern 2: Look in first few lines for date indicators
                    if date_text == "N/A":
                        lines = [l.strip() for l in post_text.split('\n') if l.strip()]
                        for i, line in enumerate(lines[:10]):
                            # Check for common date patterns
                            if any(x in line.lower() for x in ['ago', 'today', 'yesterday', 'week', 'month', 'year']):
                                date_text = line[:40]
                                break
                            # Check for time format (e.g., "12h", "2w")
                            elif re.match(r'^\d+[hdwmo]', line, re.IGNORECASE):
                                date_text = line[:20]
                                break
                    
                    # Pattern 3: Try to extract any timestamp-like content
                    if date_text == "N/A":
                        # Look for patterns with specific delimiters
                        timestamp_match = re.search(r'•\s*(.+?)\s*•', post_text)
                        if timestamp_match:
                            potential_date = timestamp_match.group(1).strip()
                            if len(potential_date) < 50 and any(c.isdigit() for c in potential_date):
                                date_text = potential_date
                except:
                    pass
            
            # Try to extract post URL from data attributes and links
            post_url = "N/A"
            post_id = "N/A"
            try:
                # Method 1: Try to find post URN from data-urn attribute (most reliable)
                urn_elems = post_elem.find_elements(By.XPATH, ".//*[@data-urn]")
                for urn_elem in urn_elems:
                    urn = urn_elem.get_attribute("data-urn")
                    if urn and ('activity' in urn or 'ugcPost' in urn or 'share' in urn):
                        # Clean URN format: urn:li:activity:7279513886758723584
                        post_url = f"https://www.linkedin.com/feed/update/{urn}/"
                        # Extract numeric ID from URN
                        id_match = re.search(r'(\d{19})', urn)
                        if id_match:
                            post_id = id_match.group(1)
                        break
                
                # Method 2: Look for timestamp/date links
                if post_url == "N/A":
                    time_links = post_elem.find_elements(By.XPATH, ".//a[contains(@href, '/feed/update/') or contains(@href, '/posts/') or contains(@class, 'app-aware-link')]")
                    for link in time_links:
                        href = link.get_attribute("href")
                        if href and ('/feed/update/' in href or '/posts/' in href):
                            post_url = href.split('?')[0]
                            # Extract numeric ID
                            id_match = re.search(r'(\d{19})', href)
                            if id_match:
                                post_id = id_match.group(1)
                            break
                
                # Method 3: Search in outer HTML for URN patterns
                if post_url == "N/A":
                    try:
                        outer_html = post_elem.get_attribute('outerHTML')
                        if outer_html:
                            # Look for URN patterns in HTML
                            urn_matches = re.findall(r'urn:li:(activity|ugcPost|share):(\d{19})', outer_html)
                            if urn_matches:
                                urn_type, numeric_id = urn_matches[0]
                                post_id = numeric_id
                                post_url = f"https://www.linkedin.com/feed/update/urn:li:{urn_type}:{post_id}/"
                    except:
                        pass
                
                # Method 4: Look for any link containing numeric ID
                if post_url == "N/A":
                    all_links = post_elem.find_elements(By.TAG_NAME, "a")
                    for link in all_links:
                        href = link.get_attribute("href") or ""
                        # Search for 19-digit LinkedIn post IDs
                        id_match = re.search(r'(\d{19})', href)
                        if id_match and ('linkedin.com' in href):
                            post_id = id_match.group(1)
                            post_url = f"https://www.linkedin.com/feed/update/urn:li:activity:{post_id}/"
                            break
            except Exception as e:
                pass
            
            # Extract reactions - try multiple methods
            reactions = 0
            try:
                # Get all text from post area that contains metrics
                post_text = post_elem.text
                
                # Method 1: Look for patterns like "1,234 reactions" or just numbers
                reaction_match = re.search(r'([\d,]+)\s*(?:reactions?|likes?)', post_text, re.IGNORECASE)
                if reaction_match:
                    reactions = int(reaction_match.group(1).replace(',', ''))
                
                # Method 2: Try aria-label approach
                if reactions == 0:
                    reaction_elems = post_elem.find_elements(By.XPATH, ".//*[contains(@aria-label, 'reaction')]")
                    if not reaction_elems:
                        reaction_elems = post_elem.find_elements(By.XPATH, ".//*[contains(@aria-label, 'like')]")
                    
                    if reaction_elems:
                        for elem in reaction_elems:
                            text = elem.get_attribute('aria-label') or elem.text
                            numbers = re.findall(r'[\d,]+', text)
                            if numbers:
                                reactions = int(numbers[0].replace(',', ''))
                                break
            except:
                pass
            
            # Extract comments - try multiple methods
            comments = 0
            try:
                post_text = post_elem.text
                
                # Method 1: Look for "X comments" pattern
                comment_match = re.search(r'([\d,]+)\s+(?:comments?)', post_text, re.IGNORECASE)
                if comment_match:
                    comments = int(comment_match.group(1).replace(',', ''))
                
                # Method 2: aria-label approach
                if comments == 0:
                    comment_elems = post_elem.find_elements(By.XPATH, ".//*[contains(@aria-label, 'comment')]")
                    if comment_elems:
                        for elem in comment_elems:
                            text = elem.get_attribute('aria-label') or elem.text
                            numbers = re.findall(r'[\d,]+', text)
                            if numbers:
                                comments = int(numbers[0].replace(',', ''))
                                break
            except:
                pass
            
            # Extract shares - try multiple methods
            shares = 0
            try:
                post_text = post_elem.text
                
                # Method 1: Look for "X shares" or "X reposts" pattern
                share_match = re.search(r'([\d,]+)\s+(?:shares?|reposts?)', post_text, re.IGNORECASE)
                if share_match:
                    shares = int(share_match.group(1).replace(',', ''))
                
                # Method 2: Look for share count in aria-label
                if shares == 0:
                    share_elems = post_elem.find_elements(By.XPATH, ".//*[contains(@aria-label, 'share') or contains(@aria-label, 'repost')]")
                    if share_elems:
                        for elem in share_elems:
                            text = elem.get_attribute('aria-label') or elem.text
                            numbers = re.findall(r'[\d,]+', text)
                            if numbers:
                                shares = int(numbers[0].replace(',', ''))
                                break
                
                # Method 3: Look for share button with count
                if shares == 0:
                    share_buttons = post_elem.find_elements(By.CSS_SELECTOR, "button[aria-label*='share'], button[aria-label*='Share']")
                    for btn in share_buttons:
                        aria_label = btn.get_attribute('aria-label') or ''
                        numbers = re.findall(r'[\d,]+', aria_label)
                        if numbers:
                            shares = int(numbers[0].replace(',', ''))
                            break
            except:
                pass
            
            # Detect media and extract URLs
            has_image = len(post_elem.find_elements(By.CSS_SELECTOR, "img")) > 0
            has_video = len(post_elem.find_elements(By.CSS_SELECTOR, "video")) > 0
            
            # Extract content/media URL
            content_url = "N/A"
            try:
                # Try to find video source first
                if has_video:
                    video_elems = post_elem.find_elements(By.CSS_SELECTOR, "video")
                    for video in video_elems:
                        video_src = video.get_attribute("src") or video.get_attribute("poster")
                        if video_src and 'http' in video_src:
                            content_url = video_src
                            break
                    
                    # Try source tags inside video
                    if content_url == "N/A":
                        source_elems = post_elem.find_elements(By.CSS_SELECTOR, "video source")
                        for source in source_elems:
                            src = source.get_attribute("src")
                            if src and 'http' in src:
                                content_url = src
                                break
                
                # Try to find image source
                elif has_image:
                    img_elems = post_elem.find_elements(By.CSS_SELECTOR, "img")
                    for img in img_elems:
                        img_src = img.get_attribute("src")
                        # Filter out profile pictures, icons, and small images
                        if img_src and 'http' in img_src and not any(x in img_src.lower() for x in ['profile', 'icon', 'emoji', 'avatar']):
                            # Check if it's a reasonable size image (not a tiny icon)
                            try:
                                width = img.get_attribute("width")
                                height = img.get_attribute("height")
                                if width and height:
                                    if int(width) > 100 and int(height) > 100:
                                        content_url = img_src
                                        break
                                else:
                                    # If no size info, take it
                                    content_url = img_src
                                    break
                            except:
                                content_url = img_src
                                break
                
                # Try to find document/attachment links
                if content_url == "N/A":
                    doc_links = post_elem.find_elements(By.XPATH, ".//a[contains(@href, '/document/') or contains(@href, '.pdf') or contains(@href, '/file/')]")
                    if doc_links:
                        doc_url = doc_links[0].get_attribute("href")
                        if doc_url:
                            content_url = doc_url
            except Exception as e:
                pass
            
            content_type = "Video" if has_video else "Image" if has_image else "Text Only"
            
            # Extract hashtags using multiple sources (text + HTML links)
            hashtags = self.extract_hashtags(
                content_text,
                # Use only the content block HTML so we don't pick up unrelated hashtags
                html_text=content_html,
                extra_texts=[content_url] if content_url and content_url != "N/A" else None,
            )
            hashtags_str = ','.join(hashtags) if hashtags else 'None'
            
            # Metrics
            char_count = len(content_text) if content_text != "N/A" else 0
            word_count = len(content_text.split()) if content_text != "N/A" else 0
            total_engagement = reactions + comments + shares
            
            followers = profile_data.get('profile_followers', 0)
            # Handle followers that might be "N/A" string
            if isinstance(followers, str):
                followers = 0
            if isinstance(followers, int) and followers > 0:
                engagement_rate = (total_engagement / followers) * 100
            else:
                engagement_rate = 0
            
            # Clean up date text - if it's too long, it's probably grabbed too much
            if date_text != "N/A" and len(date_text) > 50:
                # Try to extract just the time part (e.g., "16h", "2d", "1w")
                time_match = re.search(r'\d+[hdwmo]\b', date_text)
                if time_match:
                    date_text = time_match.group()
                else:
                    # Take first line only
                    date_text = date_text.split('\n')[0][:50]
            
            # Convert relative date to actual date
            if date_text != "N/A":
                date_text = self.convert_relative_date_to_actual(date_text)
            
            # Build post data
            post_data = {
                'Profile Username': profile_data['profile_username'],
                'Profile Name': profile_data['profile_full_name'],
                'Profile Headline': profile_data['profile_headline'],
                'Profile Location': profile_data['profile_location'],
                'Profile URL': profile_data['profile_url'],
                'Profile Followers': profile_data['profile_followers'],
                'Profile Summary': profile_data['profile_summary'],
                'Post Position': position,
                'Post URL': post_url,
                'Post ID': post_id,
                'Post Date': date_text,
                'Post Content (Full)': content_text,
                'Content URL': content_url,
                'Post Word Count': word_count,
                'Content Type': content_type,
                'Has Document': 'No',
                'Hashtags': hashtags_str,
                'Hashtag Count': len(hashtags),
                'Post Reactions (Likes)': reactions,
                'Post Comments': comments,
                'Post Shares': shares,
                'Total Engagement': total_engagement,
                'Engagement Rate (%)': f"{engagement_rate:.2f}%",
                'Comment/Reaction Ratio': f"{(comments/reactions):.3f}" if reactions > 0 else "0",
                'Share/Reaction Ratio': f"{(shares/reactions):.3f}" if reactions > 0 else "0",
                'Scraped At': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            }
            
            # VALIDATION: Only return posts with actual content
            # A valid post should have meaningful text content (not just "N/A")
            if content_text == "N/A" or len(content_text.strip()) < 10:
                print(f"    ⚠️  Skipping post {position} - no valid content found")
                return None
            
            # Also skip if the content is just a username or single word
            if word_count < 3:
                print(f"    ⚠️  Skipping post {position} - content too short ({word_count} words)")
                return None
            
            return post_data
            
        except Exception as e:
            print(f"    Error in extract_single_post: {str(e)}")
            return None
    
    def scrape_multiple_profiles(self, profile_list: list, max_posts: int = 15):
        """Scrape multiple profiles from URLs or usernames"""
        
        print(f"\n🎯 Starting multi-profile scraping")
        print(f"   Profiles to scrape: {len(profile_list)}")
        print(f"   Posts per profile: {max_posts}")
        
        success_count = 0
        failed_profiles = []
        
        for idx, profile_url in enumerate(profile_list, 1):
            print(f"\n[{idx}/{len(profile_list)}] Processing: {profile_url}")
            
            if self.scrape_profile(profile_url, max_posts=max_posts):
                success_count += 1
            else:
                failed_profiles.append(profile_url)
            
            if idx < len(profile_list):
                print(f"⏳ Waiting 3 seconds before next profile...")
                time.sleep(3)
        
        print(f"\n{'='*70}")
        print(f"✅ Successfully scraped: {success_count}/{len(profile_list)} profiles")
        if failed_profiles:
            print(f"❌ Failed profiles: {', '.join(failed_profiles)}")
        print(f"📊 Total posts collected: {len(self.all_data)}")
        print(f"{'='*70}")
        
        return success_count > 0
    
    def extract_post_data(self, post: dict, profile_data: dict, position: int) -> dict:
        """Extract post data and combine with profile data"""
        
        try:
            post_urn = post.get('urn', 'N/A')
            post_id = post_urn.split(':')[-1] if ':' in post_urn else 'N/A'
            post_url = f"https://www.linkedin.com/feed/update/{post_urn}/" if post_urn != 'N/A' else 'N/A'
            
            commentary = post.get('commentary', '')
            if isinstance(commentary, dict):
                content_text = commentary.get('text', 'N/A')
            else:
                content_text = str(commentary) if commentary else 'N/A'
            
            created_time = post.get('created', {}).get('time', 0)
            if created_time:
                date_posted = datetime.fromtimestamp(created_time / 1000).strftime('%Y-%m-%d %H:%M:%S')
            else:
                date_posted = 'N/A'
            
            total_reactions = post.get('numLikes', 0)
            total_comments = post.get('numComments', 0)
            total_shares = post.get('numShares', 0)
            
            content_type = self.detect_content_type(post)
            char_count = len(content_text) if content_text != 'N/A' else 0
            word_count = len(content_text.split()) if content_text != 'N/A' else 0
            
            import re
            hashtags = re.findall(r'#\w+', content_text)
            
            followers = profile_data.get('profile_followers', 'N/A')
            if isinstance(followers, int) and followers > 0:
                total_engagement = total_reactions + total_comments + total_shares
                engagement_rate = (total_engagement / followers) * 100
            else:
                engagement_rate = 0
            
            has_image = 'image' in str(post.get('content', {})).lower()
            has_video = 'video' in str(post.get('content', {})).lower()
            has_document = 'document' in str(post.get('content', {})).lower()
            
            combined_data = {
                'Profile Username': profile_data['profile_username'],
                'Profile Name': profile_data['profile_full_name'],
                'Profile Headline': profile_data['profile_headline'],
                'Profile Location': profile_data['profile_location'],
                'Profile URL': profile_data['profile_url'],
                'Profile Followers': profile_data['profile_followers'],
                'Profile Summary': profile_data['profile_summary'],
                'Post Position': position,
                'Post URL': post_url,
                'Post ID': post_id,
                'Post Date': date_posted,
                'Post Content (Full)': content_text,
                'Content URL': 'N/A',
                'Post Word Count': word_count,
                'Content Type': content_type,
                'Has Document': 'Yes' if has_document else 'No',
                'Hashtags': ', '.join(hashtags) if hashtags else 'None',
                'Hashtag Count': len(hashtags),
                'Post Reactions (Likes)': total_reactions,
                'Post Comments': total_comments,
                'Post Shares': total_shares,
                'Total Engagement': total_reactions + total_comments + total_shares,
                'Engagement Rate (%)': f"{engagement_rate:.2f}%",
                'Comment/Reaction Ratio': f"{(total_comments/total_reactions):.3f}" if total_reactions > 0 else "0",
                'Share/Reaction Ratio': f"{(total_shares/total_reactions*100):.2f}%" if total_reactions > 0 else "0%",
                'Scraped At': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            }
            
            return combined_data
            
        except Exception as e:
            print(f"    Error extracting post: {str(e)}")
            return None
    
    def detect_content_type(self, post: dict) -> str:
        """Detect content type from post data"""
        try:
            content_str = str(post.get('content', {})).lower()
            if 'video' in content_str:
                return 'Video'
            elif 'document' in content_str:
                return 'Document/Carousel'
            elif 'article' in content_str:
                return 'Article/Link'
            elif 'image' in content_str:
                return 'Image'
            elif 'poll' in content_str:
                return 'Poll'
            else:
                return 'Text Only'
        except:
            return 'Unknown'
    
    def scrape_multiple_profiles(self, profile_list: list, max_posts: int = 15):
        """Scrape multiple profiles in one run"""
        print(f"\n🎯 Starting multi-profile scraping")
        print(f"   Profiles to scrape: {len(profile_list)}")
        print(f"   Posts per profile: {max_posts}")
        
        success_count = 0
        failed_profiles = []
        
        for idx, username in enumerate(profile_list, 1):
            print(f"\n[{idx}/{len(profile_list)}] Processing @{username}")
            
            if self.scrape_profile(username, max_posts=max_posts):
                success_count += 1
            else:
                failed_profiles.append(username)
            
            if idx < len(profile_list):
                print(f"⏳ Waiting 3 seconds before next profile...")
                time.sleep(3)
        
        print(f"\n{'='*70}")
        print(f"✅ Successfully scraped: {success_count}/{len(profile_list)} profiles")
        if failed_profiles:
            print(f"❌ Failed profiles: {', '.join(failed_profiles)}")
        print(f"📊 Total posts collected: {len(self.all_data)}")
        print(f"{'='*70}")
        
        return success_count > 0
    
    
    def save_to_excel(self, filename: str = None):
        """Save all data to single Excel sheet"""
        
        if not filename:
            filename = f"linkedin_multi_profile_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        if not self.all_data:
            print("❌ No data to save!")
            return False
        
        print(f"\n💾 Saving to {filename}...")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "All Posts"
        
        headers = list(self.all_data[0].keys())
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="1F1F1F", end_color="1F1F1F", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for post in self.all_data:
            row = [post.get(header, 'N/A') for header in headers]
            ws.append(row)
        
        column_widths = {
            'A': 15, 'B': 20, 'C': 30, 'D': 15, 'E': 35, 'F': 12, 'G': 12, 'H': 30,
            'I': 10, 'J': 35, 'K': 20, 'L': 18, 'M': 40, 'N': 50,
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        for column in ws.columns:
            column_letter = column[0].column_letter
            if column_letter not in column_widths:
                max_length = 0
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        ws.freeze_panes = 'C2'
        wb.save(filename)
        
        print(f"✅ Saved: {filename}")
        print(f"📊 Total rows: {len(self.all_data)} posts")
        print(f"📋 Columns: {len(headers)}")
        
        print(f"\n📈 Summary by Profile:")
        profile_counts = {}
        for post in self.all_data:
            username = post['Profile Username']
            profile_counts[username] = profile_counts.get(username, 0) + 1
        
        for username, count in profile_counts.items():
            print(f"   @{username}: {count} posts")
        
        return True
    
    def close(self):
        """Close browser"""
        if self.driver:
            self.driver.quit()
            print("🔒 Browser closed")


def get_cookies_instructions():
    """Display instructions for using the scraper"""
    print("""
╔══════════════════════════════════════════════════════════════════════╗
║           LinkedIn Scraper - Headless Browser Mode                   ║
╚══════════════════════════════════════════════════════════════════════╝

✅ HOW IT WORKS:
──────────────────────────────────────────────────────────────────────
1. Browser opens INVISIBLY in background (headless mode)
2. Logs in with your credentials automatically
3. Scrapes profiles and posts
4. Browser closes when done
5. No window stays on screen - everything happens in terminal!

🔐 CREDENTIALS:
──────────────────────────────────────────────────────────────────────
Update these in the script:
   LINKEDIN_EMAIL = "your-email@gmail.com"
   LINKEDIN_PASSWORD = "your-password"

📋 PROFILES:
──────────────────────────────────────────────────────────────────────
Add profiles to scrape:
   PROFILES_TO_SCRAPE = [
       "profile1",
       "profile2",
       ...
   ]

🎯 RUN THE SCRIPT:
──────────────────────────────────────────────────────────────────────
   python linkedin_scraper.py

Everything happens in the background - no visible browser!
    """)


def main():
    """Main execution"""
    
    print("=" * 70)
    print("🔵 LinkedIn Multi-Profile Scraper (Headless Browser)")
    print("=" * 70)
    
    # ========================================
    # CONFIGURE THIS SECTION
    # ========================================
    
    LINKEDIN_EMAIL = "mtarate2004@gmail.com"
    LINKEDIN_PASSWORD = "Janki@200904"
    
    # Add LinkedIn profile URLs here (personal or company pages)
    PROFILES_TO_SCRAPE = [
        "https://www.linkedin.com/company/ikfdigital/",
        "https://www.linkedin.com/company/amura-marketing-technologies-pvt-ltd/",
        "https://www.linkedin.com/company/communicate-india/posts/?feedView=all&viewAsMember=true"  
    ]
    
    
    MAX_POSTS_PER_PROFILE = 15
    
    # ========================================
    
    print(f"\n⚙️  Configuration:")
    print(f"   Email: {LINKEDIN_EMAIL}")
    print(f"   Profiles to scrape: {len(PROFILES_TO_SCRAPE)}")
    print(f"   Posts per profile: {MAX_POSTS_PER_PROFILE}")
    print(f"\n📋 Target Profiles:")
    for i, url in enumerate(PROFILES_TO_SCRAPE, 1):
        print(f"   {i}. {url}")
    
    scraper = LinkedInMultiProfileScraper(
        email=LINKEDIN_EMAIL,
        password=LINKEDIN_PASSWORD
    )
    
    try:
        # Setup headless browser
        scraper.setup_driver()
        
        # Login
        if not scraper.login():
            print("\n❌ Login failed")
            return
        
        # Scrape profiles
        if scraper.scrape_multiple_profiles(PROFILES_TO_SCRAPE, max_posts=MAX_POSTS_PER_PROFILE):
            scraper.save_to_excel()
            
            if scraper.all_data:
                total_reactions = sum(post['Post Reactions (Likes)'] for post in scraper.all_data)
                total_comments = sum(post['Post Comments'] for post in scraper.all_data)
                total_engagement = sum(post['Total Engagement'] for post in scraper.all_data)
                
                print(f"\n✅ SUCCESS!")
                print(f"   Total Profiles: {len(PROFILES_TO_SCRAPE)}")
                print(f"   Total Posts: {len(scraper.all_data)}")
                print(f"   Total Reactions: {total_reactions:,}")
                print(f"   Total Comments: {total_comments:,}")
                print(f"   Total Engagement: {total_engagement:,}")
        
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
    finally:
        # Always close browser
        scraper.close()
    
    print("\n✅ Done!")


if __name__ == "__main__":
    main()